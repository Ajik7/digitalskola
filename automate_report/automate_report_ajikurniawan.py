#Membuat Automate Reporting ( Table & Grafik ), dan dikirim ke discord
#Aji Kurniawan

from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import *
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

input_file = 'input_data\supermarket_sales.xlsx'
output_file= "data_output/report_penjualan_2019.xlsx"
webhook_url='https://discord.com/api/webhooks/1026741978718814259/iIvIpko9qE-q6072J9DEXyPXFkUdYzf8sh_efSHH7g70Rwxy_m8L7nHeDCfdxPhFdcwp'
#PART 1 LOAD DATASET
df=pd.read_excel(input_file)

df=df.pivot_table(index='Gender',
                    columns='Product line',
                    values='Total', 
                    aggfunc='sum').round()


print('Save dataframe to excel....')
df.to_excel(output_file, sheet_name='Report',startrow=4)
print('Save dataframe to excel done...')

#PART 2 GRAFIK

wb= load_workbook(output_file)
wb.active=wb['Report']

min_column=wb.active.min_column
max_column=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

barchart=BarChart()
data=Reference(wb.active,
                min_col=min_column+1,
                max_col=max_column,
                min_row=min_row,
                max_row=max_row)

categories=Reference(wb.active,
                        min_col=min_column,
                        max_col=max_column,
                        min_row=min_row+1,
                        max_row=max_row)

barchart.add_data(data,titles_from_data=True)
barchart.set_categories(categories)

wb.active.add_chart(barchart,'B12')
barchart.title='Sales Berdasarkan Produk'
barchart.style=2
#wb.save(output_file)

#Total Penjualan
import string
alphabet=list(string.ascii_uppercase)
alphabet_excel=alphabet[:max_column]
for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}']=f'=SUM({i}{min_row+1}:{max_row})'
        wb.active[f'{i}{max_row+1}'].style='Currency'
wb.active[f'{alphabet_excel[0]}{max_row+1}']='Total'
wb.active['A1']='Sales Report'
wb.active['A2']='2019'
wb.active['A1'].font=Font('Arial', bold=True,size=20)
wb.active['A2'].font=Font('Arial', bold=True,size=20)
wb.save(output_file)

#Send to Discord
def send_to_discord():
    import requests
    import discord
    from discord import SyncWebhook
    webhook=SyncWebhook.from_url(webhook_url)
    with open(file=output_file, mode='rb') as file:
        excel_file=discord.File(file)

    webhook.send('This is an automate report',username='Sales Bot', file=excel_file)

send_to_discord()

